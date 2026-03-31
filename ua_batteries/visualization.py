"""Visualization, HTML export, and Excel export for optimization results."""

from __future__ import annotations

import os
import tempfile
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ua_batteries.config import CAPACITY, MAX_BUYS, MAX_SELLS, POWER
from ua_batteries.main import add_optimization_to_dataframe

GREEN_FILL = PatternFill(fill_type="solid", fgColor="90EE90")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFB6C1")
PROFIT_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")
DATE_FILL = PatternFill(fill_type="solid", fgColor="E8F5E9")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4CAF50")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E3F2FD")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="0F4C81")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def create_optimization_visualization(
    df,
    max_buys: float = MAX_BUYS,
    max_sells: float = MAX_SELLS,
    capacity: float = CAPACITY,
    power: float = POWER,
) -> pd.DataFrame:
    """Create a visualization dataframe of the buy/sell optimization strategy."""
    if "Buy_hours" not in df.columns:
        df = add_optimization_to_dataframe(
            df,
            max_buys=max_buys,
            max_sells=max_sells,
            capacity=capacity,
            power=power,
        )

    viz_df = pd.DataFrame(index=df.index)
    for hour in range(1, 25):
        viz_df[f"Hour {hour}"] = ""

    for idx, row in df.iterrows():
        buy_hours = row["Buy_hours"]
        sell_hours = row["Sell_hours"]

        if buy_hours:
            for buy_info in buy_hours.values():
                hour = buy_info["hour"]
                amount = buy_info["buy_amount"]
                viz_df.at[idx, f"Hour {hour + 1}"] = f"BUY\n{amount:.2f}MWh"

        if sell_hours:
            for sell_info in sell_hours.values():
                hour = sell_info["hour"]
                amount = sell_info["sell_amount"]
                viz_df.at[idx, f"Hour {hour + 1}"] = f"SELL\n{amount:.2f}MWh"

    viz_df["Total Profit"] = df["Total_profit"].values
    return viz_df


def style_visualization(viz_df: pd.DataFrame, df: pd.DataFrame):
    """Apply styling to the visualization dataframe."""
    def color_cells(val: Any, col_name: str) -> str:
        if col_name == "Total Profit":
            return "font-weight: bold"

        if not val or val == "":
            return ""

        if "BUY" in str(val):
            return "background-color: #90EE90; color: black; font-weight: bold"
        if "SELL" in str(val):
            return "background-color: #FFB6C1; color: black; font-weight: bold"
        return ""

    styled = viz_df.style
    for col in viz_df.columns:
        styled = styled.map(lambda v, current_col=col: color_cells(v, current_col), subset=[col])
    return styled


def display_visualization(
    df,
    max_buys: float = MAX_BUYS,
    max_sells: float = MAX_SELLS,
    capacity: float = CAPACITY,
    power: float = POWER,
    return_styled: bool = False,
):
    """Create and display the optimization visualization."""
    optimized_df = add_optimization_to_dataframe(
        df,
        max_buys=max_buys,
        max_sells=max_sells,
        capacity=capacity,
        power=power,
    )
    viz_df = create_optimization_visualization(
        optimized_df,
        max_buys=max_buys,
        max_sells=max_sells,
        capacity=capacity,
        power=power,
    )
    if return_styled:
        return style_visualization(viz_df, optimized_df)
    return viz_df


def export_to_html(viz_df: pd.DataFrame, original_df: pd.DataFrame, title: str = "Energy Trading Optimization", open_browser: bool = True) -> str:
    """Export visualization to styled HTML file with prices and optionally open in browser."""
    html_string = f"""
    <html>
    <head>
        <title>{title}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }}
            h1 {{
                color: #333;
                text-align: center;
            }}
            .container {{
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                overflow-x: auto;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 10px;
                text-align: center;
                font-size: 11px;
            }}
            th {{
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
            .buy {{
                background-color: #90EE90;
                font-weight: bold;
                color: darkgreen;
            }}
            .sell {{
                background-color: #FFB6C1;
                font-weight: bold;
                color: darkred;
            }}
            .profit {{
                background-color: #FFF9C4;
                font-weight: bold;
                color: #F57F17;
            }}
            .date-col {{
                background-color: #E8F5E9;
                font-weight: bold;
                text-align: left;
            }}
            .price {{
                font-size: 10px;
                display: block;
                color: #666;
            }}
            .amount {{
                font-size: 10px;
                font-weight: bold;
                display: block;
                white-space: pre-line;
            }}
            .summary {{
                margin-top: 20px;
                padding: 15px;
                background-color: #E3F2FD;
                border-left: 4px solid #2196F3;
                border-radius: 4px;
            }}
            .summary h2 {{
                margin-top: 0;
                color: #1976D2;
            }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <div class="container">
            <table>
                <tr>
                    <th>Date</th>
    """

    for hour in range(1, 25):
        html_string += f"<th>Hour {hour}</th>"

    html_string += "<th>Total Profit</th></tr>"

    for i, (_, row) in enumerate(viz_df.iterrows()):
        date_value = original_df.index[i]
        html_string += f"<tr><td class='date-col'>{date_value}</td>"
        original_row = original_df.iloc[i]

        for hour in range(1, 25):
            price = original_row[str(hour)]
            cell_value = row[f"Hour {hour}"]

            if "BUY" in str(cell_value):
                html_string += f"<td class='buy'><span class='price'>{price:.2f}</span><span class='amount'>{cell_value}</span></td>"
            elif "SELL" in str(cell_value):
                html_string += f"<td class='sell'><span class='price'>{price:.2f}</span><span class='amount'>{cell_value}</span></td>"
            else:
                html_string += f"<td><span class='price'>{price:.2f}</span></td>"

        profit = row["Total Profit"]
        html_string += f"<td class='profit'>{profit:,.2f}</td></tr>"

    total_profit = viz_df["Total Profit"].sum()
    html_string += f"""
            </table>
        </div>
        <div class="summary">
            <h2>Summary</h2>
            <p><strong>Total Profit for Period:</strong> {total_profit:,.2f} (UAH)</p>
            <p><strong>Number of Days:</strong> {len(viz_df)}</p>
            <p><strong>Average Daily Profit:</strong> {total_profit / len(viz_df):,.2f} (UAH)</p>
        </div>
    </body>
    </html>
    """

    temp_dir = tempfile.gettempdir()
    html_file = os.path.join(temp_dir, "energy_optimization_viz.html")
    with open(html_file, "w", encoding="utf-8") as file_obj:
        file_obj.write(html_string)

    print(f"HTML visualization saved to: {html_file}")
    if open_browser:
        webbrowser.open(f"file://{html_file}")
        print("Opening in default browser...")

    return html_file


def _base_output_path(output_path: str | None, suffix: str) -> str:
    """Return an output path, using a temp file when no path is supplied."""
    if output_path:
        return output_path
    temp_dir = tempfile.gettempdir()
    return os.path.join(temp_dir, f"energy_optimization_viz{suffix}")


def export_to_excel(
    viz_df: pd.DataFrame,
    original_df: pd.DataFrame,
    title: str = "Energy Trading Optimization",
    output_path: str | None = None,
) -> str:
    """Export the optimization calendar to an Excel workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Optimization"

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B4"

    title_row = 1
    header_row = 3
    start_data_row = 4

    last_col = 26
    worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_col)
    title_cell = worksheet.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(color="FFFFFF", bold=True, size=15)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER
    worksheet.row_dimensions[title_row].height = 24

    worksheet.cell(row=header_row, column=1, value="Date")
    for hour in range(1, 25):
        worksheet.cell(row=header_row, column=hour + 1, value=f"Hour {hour}")
    worksheet.cell(row=header_row, column=26, value="Total Profit")

    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=header_row, column=column)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for idx, (_, viz_row) in enumerate(viz_df.iterrows(), start=start_data_row):
        original_row = original_df.iloc[idx - start_data_row]
        worksheet.cell(row=idx, column=1, value=str(original_df.index[idx - start_data_row]))
        date_cell = worksheet.cell(row=idx, column=1)
        date_cell.fill = DATE_FILL
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        date_cell.border = THIN_BORDER

        for hour in range(1, 25):
            price = float(original_row[str(hour)])
            cell_value = str(viz_row[f"Hour {hour}"]).strip()
            text_value = f"{price:.2f}"
            if cell_value:
                text_value = f"{price:.2f}\n{cell_value}"

            cell = worksheet.cell(row=idx, column=hour + 1, value=text_value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            cell.font = Font(size=10)

            if "BUY" in cell_value:
                cell.fill = GREEN_FILL
                cell.font = Font(size=10, bold=True, color="006400")
            elif "SELL" in cell_value:
                cell.fill = RED_FILL
                cell.font = Font(size=10, bold=True, color="8B0000")

        profit_cell = worksheet.cell(row=idx, column=26, value=float(viz_row["Total Profit"]))
        profit_cell.number_format = '#,##0.00'
        profit_cell.fill = PROFIT_FILL
        profit_cell.font = Font(bold=True, color="F57F17")
        profit_cell.alignment = Alignment(horizontal="center", vertical="center")
        profit_cell.border = THIN_BORDER

    for row_idx in range(start_data_row, start_data_row + len(viz_df)):
        worksheet.row_dimensions[row_idx].height = 36

    worksheet.column_dimensions["A"].width = 24
    for col_idx in range(2, 26):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 12
    worksheet.column_dimensions["Z"].width = 14
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 12

    summary_start = start_data_row + len(viz_df) + 2
    worksheet.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=5)
    summary_title = worksheet.cell(row=summary_start, column=1, value="Summary")
    summary_title.fill = SUMMARY_FILL
    summary_title.font = Font(bold=True, color="1976D2")
    summary_title.alignment = Alignment(horizontal="left", vertical="center")
    summary_title.border = THIN_BORDER

    data_start = start_data_row
    data_end = start_data_row + len(viz_df) - 1
    profit_col = "Z"

    summary_rows = [
        ("Total Profit for Period", f"=SUM({profit_col}{data_start}:{profit_col}{data_end})", "UAH"),
        ("Number of Days", f"=COUNTA(A{data_start}:A{data_end})", ""),
        (
            "Average Daily Profit",
            f"=IF(B{summary_start + 2}=0,0,B{summary_start + 1}/B{summary_start + 2})",
            "UAH",
        ),
    ]

    for offset, (label, formula_or_value, unit) in enumerate(summary_rows, start=1):
        label_cell = worksheet.cell(row=summary_start + offset, column=1, value=label)
        value_cell = worksheet.cell(row=summary_start + offset, column=2, value=formula_or_value)
        unit_cell = worksheet.cell(row=summary_start + offset, column=3, value=unit)

        for current_cell in (label_cell, value_cell, unit_cell):
            current_cell.border = THIN_BORDER

        label_cell.fill = SUMMARY_FILL
        label_cell.font = Font(bold=True)
        value_cell.font = Font(bold=True)
        value_cell.number_format = '#,##0.00'
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        unit_cell.font = Font(italic=True, color="666666")

    output_file = _base_output_path(output_path, ".xlsx")
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    return output_file


if __name__ == "__main__":
    from ua_batteries.config import REQUEST_DAY
    from ua_batteries.utils.get_file import get_file

    df = get_file()
    print("\nRunning optimization (this may take a moment)...")
    viz = create_optimization_visualization(df, max_buys=MAX_BUYS, max_sells=MAX_SELLS, capacity=CAPACITY, power=POWER)

    date = datetime.strptime(REQUEST_DAY, "%m.%Y").strftime("%B %Y")
    html_path = export_to_html(viz, df, title=f"Energy Trading Optimization - {date}", open_browser=True)
    excel_path = export_to_excel(viz, df, title=f"Energy Trading Optimization - {date}", output_path=f"file_downloads/Energy Trading Optimization - {date}.xlsx")

    print("\nVisualization complete!")
    print(f"Total profit for period: {viz['Total Profit'].sum():.2f}")
    print(f"HTML: {html_path}")
    print(f"Excel: {excel_path}")
